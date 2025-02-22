from datetime import datetime
from email import message
from math import e
from pickle import EXT2
from re import sub
from typing import Dict, List, Literal, Union
from venv import create
from fastapi import APIRouter, Body, Depends, HTTPException, Query,Header
from bson import ObjectId
from seaborn import dark_palette
from sympy import sec
from models.AdminModel import ExamTimetable,UpdateCRRequest,YearAssignment,TodayClassesRequest,ClassAttendanceRequest
from numpy import number
from models.FacultyModel import Attendance, Faculty
from models.StudentModel import Branch, Student
from pydantic import BaseModel
from .faculty import calculate_percentage
from db import database
from .faculty import calculate_percentage
from db.database import db,admin, faculty, student,db2,db3 ,db4,db5 # Assuming these collections exist
from routes import auth  # For password hashing and verification
from fastapi.responses import StreamingResponse
import json
from openpyxl import Workbook
from openpyxl.styles import Alignment,Font,Border,Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import os
from models.AdminModel import TimeTableRequest

attendance_collections = {
        'E1': database.E1,
        'E2': database.E2, 
        'E3': database.E3,
        'E4': database.E4
    }

subjects_collections = {
    'E1':database.E1_subjects,
    'E2':database.E2_subjects,
    'E3':database.E3_subjects,
    'E4':database.E4_subjects
}

exam_timetable_collections = {
    'E1': database.E1_exam_timetable,
    'E2': database.E2_exam_timetable,
    'E3': database.E3_exam_timetable,
    'E4': database.E4_exam_timetable,
        }
student_collections = {
    'E1': database.E1_student,
    'E2': database.E2_student,
    'E3': database.E3_student,
    'E4': database.E4_student,
            }

timetable_collections ={
    "E1":database.E1_timetable,
    "E2":database.E2_timetable,
    "E3":database.E3_timetable,
    "E4":database.E4_timetable
}
router = APIRouter()

# async def check_token(token):

async def check_admin_email(email: str):
    admin_data = await admin.find_one({"email_address": email})
    if admin_data:
        return True 
    else:
        return False


@router.put("/update-year-and-semester/")
async def update_years_and_sem(request: str,user: dict = Depends(auth.get_current_user)):
    """
    Bulk updates students' year and semester based on the request type.
    Args:
        request: A string indicating whether to update "year" or "sem".
    """
    if user['role']!='admin' or (not check_admin_email(user['email_address'])):
        return {'status_code':500,'message':'you cannot access this page'}
    if request not in ["year", "sem"]:
        raise HTTPException(status_code=400, detail="Invalid request. Must be 'year' or 'sem'.")

    students = await student.find().to_list(None) # Fetch all students
    if not students:
        raise HTTPException(status_code=404, detail="No students found.")

    updates = []

    for student_data in students:
        current_year = student_data["year"]
        current_semester = student_data["semester"]

        if request == "year":
            if current_year == "E4" and current_semester == 2:
                continue  # Skip students already in the final semester
            
            new_year = {"E1": "E2", "E2": "E3", "E3": "E4"}.get(current_year, current_year)
            new_semester = 1 if current_semester == 2 else current_semester  # Reset to sem-1 if it's sem-2
            
            updates.append({
                "id_number": student_data["id_number"],
                "year": new_year,
                "semester": new_semester
            })
        
        elif request == "sem":
            if current_semester == 1:
                updates.append({
                    "id_number": student_data["id_number"],
                    "semester": 2
                })
    if not updates:
        raise HTTPException(status_code=400, detail="No students were eligible for updates.")

    # Perform bulk update
    for update in updates:
        await student.update_one(
            {"id_number": update["id_number"]},
            {"$set": {"year": update.get("year", student_data["year"]), "semester": update.get("semester", student_data["semester"])}}
        )
    
    return {"message": "Students updated successfully", "updated_students": len(updates)}


@router.put("/update-cr-id/")
async def update_cr_id(request: UpdateCRRequest):
    """
    Update the CR ID for a specific section and year in the student database.
    """
    year = request.year
    section_name = request.section_name
    cr_id = request.cr_id

    # Find the section based on year and section_name
    year_data=db4[year]
    # print(year_data)
    section = await year_data.find_one({"section_name": section_name})
    # print(section)
    if not section:
        raise HTTPException(status_code=404, detail=f"Section {section_name} in year {year} not found.")

    # Check if the student belongs to the section
    if cr_id not in section.get("students", []):
        raise HTTPException(
            status_code=403, 
            detail=f"Student {cr_id} does not belong to section {section_name} in year {year}."
        )

    # Update the CR ID
    update_result = await db4[year].update_one(
        {"_id": section["_id"]}, {"$set": {"cr_id": cr_id}}
    )

    if update_result.modified_count == 0:
        raise HTTPException(status_code=500, detail="Failed to update CR ID.")
    
    return {"message": f"CR ID for section {section_name} in year {year} updated successfully."}


async def get_subjects_for_section(
    year: str, section_name: str, db4
) -> List[str]:
    """
    Fetch the list of subjects available for a specific section in a given year.

    Args:
    - year: Year of the section (e.g., "e1", "e2").
    - section_name: Name of the section (e.g., "A", "B").
    - db: The database instance.

    Returns:
    - A list of subject names for the given section.
    """
    # Access the collection for the specified year
    year_collection = db4[year]

    # Find the section based on the section name
    section_data = await year_collection.find_one({"section_name": section_name})

    if not section_data:
        raise HTTPException(
            status_code=404, detail=f"Section {section_name} in year {year} not found."
        )

    # Extract the subject names from the section data
    subjects = [subject["subject_name"] for subject in section_data.get("subjects", [])]

    if not subjects:
        raise HTTPException(
            status_code=404,
            detail=f"No subjects found for section {section_name} in year {year}.",
        )
    return subjects


# Route to store or update the exam timetable
@router.put("/update-exam-timetable/")
async def update_exam_timetable(data: ExamTimetable,user: dict = Depends(auth.get_current_user)):
    """
    Stores or updates the exam timetable for a specific year and semester.
    """
    if user['role']!='admin' or (not check_admin_email(user['email_address'])):
        return {'status_code':500,'message':'you cannot access this page'}
    year_collection = db5[data.year]  # Access the collection for the specified year

    # Fetch the subjects for the specified section
    try:
        section_subjects = await get_subjects_for_section(data.year, 'A', db4)
    except HTTPException as e:
        raise HTTPException(status_code=e.status_code, detail=e.detail)

    # Gather all subjects listed in sem_exam
    timetable_subjects = set(data.sem_exam.keys())

    # Validate the subjects
    invalid_subjects = timetable_subjects - set(section_subjects)
    if invalid_subjects:
        raise HTTPException(
            status_code=400,
            detail=f"The following subjects in the timetable are invalid for year {data.year}: {', '.join(invalid_subjects)}"
        )

    # Check if a timetable for the given semester already exists
    existing_timetable = await year_collection.find_one({"semester": data.semester})

    if existing_timetable:
        # Update the existing timetable
        result = await year_collection.update_one(
            {"_id": existing_timetable["_id"]},
            {"$set": {
                "mid_exams": data.mid_exams,
                "sem_exam": data.sem_exam
            }}
        )
        if result.modified_count:
            return {"message": f"Timetable for semester {data.semester} in {data.year} updated successfully."}
        else:
            raise HTTPException(status_code=500, detail="Failed to update the timetable.")
    else:
        # Insert a new timetable
        result = await year_collection.insert_one({
            "semester": data.semester,
            "mid_exams": data.mid_exams,
            "sem_exam": data.sem_exam
        })
        if result.inserted_id:
            return {"message": f"Timetable for semester {data.semester} in {data.year} added successfully."}
        else:
            raise HTTPException(status_code=500, detail="Failed to add the timetable.")


@router.get('/get-student-attendance')
async def get_student_attendance(student_id: str, year: str,user: dict = Depends(auth.get_current_user)):
    """
    Get the attendance details of a student for a specific year.
    """
    if user['role']!='admin' or (not check_admin_email(user['email'])):
        return {'status_code':500,'message':'you cannot access this page'}
    # Fetch the student's attendance data for the year
    attendance_data = attendance_collections[year]
    student_attendance = await attendance_data.find_one({"id_number": student_id})
    student_details = await student.find_one({"id_number": student_id})
    
    if not student_details: 
        raise HTTPException(status_code=404, detail=f"Student {student_id} not found.")
    if not student_attendance:
        raise HTTPException(status_code=404, detail=f"Attendance data not found for student {student_id} in year {year}.")
    for field in ['_id', 'password', 'semester','is_admin']:
        student_details.pop(field, None)
    subjects_report = calculate_percentage(student_attendance)
    subjects_report.pop('total_percentage',None)
    for key,value in subjects_report.items():
        value.pop('faculty_name',None)
    return {'status_code':200,'student_details':student_details,'attendance':student_attendance['attendance_report'],'subjects_report':subjects_report}

@router.post("/update-timetable-for-faculty/")
async def update_timetable_for_faculty(assignments: YearAssignment,user: dict = Depends(auth.get_current_user)):
    """
    Incrementally update the timetable and faculty data for the given year.
    Dynamically add new subjects if they do not exist in the timetable.
    """
    
    if user['role']!='admin' or (not check_admin_email(user['email_address'])):
        return {'status_code':500,'message':'you cannot access this page'}
    year = assignments.year
    timetable_collection = db3[year]  # Access the year-specific timetable collection
    faculty_collection = db["Faculty"]  # Faculty collection

    # Fetch available subjects for the year
    year_data = db4[year]
    available_subjects = set()

    async for semester_data in year_data.find():
        available_subjects.update(
            subject["subject_name"] for subject in semester_data.get("subjects", [])
        )

    if not available_subjects:
        raise HTTPException(
            status_code=404,
            detail=f"No subjects found for year {year} in the database."
        )

    # Fetch the existing timetable document or initialize a new one
    existing_timetable_doc = await timetable_collection.find_one({"type": "assignments"})
    timetable_doc = existing_timetable_doc or {"type": "assignments", "subjects": {}}

    for subject_assignment in assignments.assignments:
        subname = subject_assignment.subname

        # Check if the subject is valid for the year
        if subname not in available_subjects:
            # If subname is new, dynamically add it to the timetable doc
            timetable_doc["subjects"].setdefault(subname, [])

        # Prepare the new data for this subject
        new_faculty_data = [
            {"faculty_username": fa.faculty_username, "sec": fa.sec}
            for fa in subject_assignment.data
        ]

        # Merge or update the timetable subject data
        existing_subject_data = timetable_doc["subjects"].get(subname, [])

        updated_subject_data = []
        for new_faculty in new_faculty_data:
            existing_match = next(
                (ef for ef in existing_subject_data if ef["faculty_username"] == new_faculty["faculty_username"]),
                None
            )
            if existing_match:
                # Merge sections if there are overlaps
                existing_match["sec"] = list(set(existing_match["sec"]) | set(new_faculty["sec"]))
                updated_subject_data.append(existing_match)
            else:
                # Add new assignment if not present
                updated_subject_data.append(new_faculty)

        # Update timetable with the merged data
        timetable_doc["subjects"][subname] = updated_subject_data

        # Update each faculty's subject list
        for fa in subject_assignment.data:
            faculty = await faculty_collection.find_one({
               'email_address':fa.faculty_username
            })

            if not faculty:
                return {'status_code':404,
                        # 'detail':f'Faculty {fa.faculty_username} not found'
                        }

            # Merge or update the faculty's subject list
            existing_subjects = faculty.get("subjects", [])
            existing_assignment = next(
                (subj for subj in existing_subjects if subj["subject_name"] == subname), None
            )

            if existing_assignment:
                # Merge sections if they overlap
                existing_assignment["sections"] = list(set(existing_assignment["sections"]) | set(fa.sec))
            else:
                # Add new assignment if not present
                existing_subjects.append({
                    "subject_name": subname,
                    "year": year,
                    "sections": fa.sec
                })

            # Update the faculty's subject list in the database
            await faculty_collection.update_one(
                {"_id": faculty["_id"]},
                {"$set": {"subjects": existing_subjects}}
            )

    # Store the updated timetable document in the database
    if existing_timetable_doc:
        # Update only if the document has changed
        await timetable_collection.update_one(
            {"_id": existing_timetable_doc["_id"]},
            {"$set": {"subjects": timetable_doc["subjects"]}}
        )
    else:
        # Insert a new document
        await timetable_collection.insert_one(timetable_doc)

    return {'status_code':200}


# # Helper function to calculate percentage
# def calculate_percentage(part: int, whole: int) -> float:
#     if whole == 0:
#         return 0.0
#     return round((part / whole) * 100, 2)
@router.post("/Today_classes")
async def admin_Today_classes(data: TodayClassesRequest,user: dict = Depends(auth.get_current_user)):
    """
    Admin dashboard to view timetable and attendance for all years and sections.
    The `date` parameter is in YYYY-MM-DD format.
    """
    # Validate date
    if user['role']!='admin' or (not check_admin_email(user['email'])):
        return {'status_code':500,'message':'you cannot access this page'}
        
    date = data.today_date
    try:
        query_date_obj = datetime.strptime(date, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid date format. Use YYYY-MM-DD.")

    day_of_week = query_date_obj.strftime("%A").lower()
    today_date_obj = datetime.now()

    overall_response = {}

    # Iterate through all years (e.g., E1, E2, E3, E4)
    year = data.year
    # Fetch timetable for the year
    timetable_collection = db3[year]
    timetable_data = await timetable_collection.find_one({})
    faculty_data = await timetable_collection.find_one({'type':'assignments'})
    if not timetable_data or day_of_week not in timetable_data:
        raise HTTPException(status_code=404, detail=f"Timetable not found for {year} on {day_of_week}.")

    day_timetable = timetable_data[day_of_week]
    year_data = {}
    # print(day_timetable)
    for section, subjects in day_timetable.items():
        year_data[section] = {}  
        for subject, periods in subjects.items():
            faculty_details_list = []

            for faculty_entry in faculty_data['subjects'].get(subject, []):
                if section in faculty_entry['sec']: 
                    faculty_email = faculty_entry['faculty_username']
                        
                    faculty_details = await database.faculty.find_one({'email_address': faculty_email})
                        
                    if faculty_details:
                        faculty_name = f"{faculty_details['first_name']} {faculty_details['last_name']}"
                        faculty_phone = faculty_details['phone_number']

                        faculty_details_list.append({
                            "faculty_name": faculty_name,
                            "faculty_phone": faculty_phone,
                            "email_address": faculty_email
                        })

            year_data[section][subject] = {
                "faculty": faculty_details_list,
                "periods": periods
            }
    overall_response[year] = year_data    
    return overall_response

@router.post('/class-attendance')
async def class_attendance(data:ClassAttendanceRequest,user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin' or (not check_admin_email(user['email'])):
        return {'status_code':500,'message':'you cannot access this page'}
    year = data.year
    section = data.section
    date = data.date
    subject = data.subject
    student_database = student_collections[year]
    attendance_collection = attendance_collections[year]
    section_datails = await student_database.find_one({'section_name':section})
    presenties = []
    absenties = []
    number_of_periods = 0
    cancelled=True
    for student in section_datails['students']:
        student_attendance = await attendance_collection.find_one({'id_number':student})
        if not subject in student_attendance['attendance_report']:
            raise HTTPException(status_code=404,detail="subject not found")
        else:
            for record in student_attendance["attendance_report"][subject]["attendance"]:
                if record["date"] == date:
                    cancelled=False
                    if record["status"] == "present":
                        presenties.append(student)
                        number_of_periods = record["number_of_periods"]
                    else:
                        absenties.append(student)
                        number_of_periods = record["number_of_periods"]
                    break
                
    if cancelled:
        return{'message':'Class is cancelled'}
    else:
        percentage = len(presenties) / len(section_datails['students']) * 100
        
        return {'presenties':presenties,'absenties':absenties,'percentage':percentage,'number_of_periods':number_of_periods}
                    
                
@router.get('/dashboard')
async def admin_dashboard(user: dict = Depends(auth.get_current_user)):      
    if user['role']!='admin' or (not check_admin_email(user['email'])):
        return {'status_code':500,'message':'you cannot access this page'}
    #all years percentage
    years = ['E1']
    overall_response = {}
    for year in years:
        year_data = []
        year_collection = attendance_collections[year]
        attendance_data = await year_collection.find({}).to_list(None)
        for record in attendance_data:
            result = calculate_percentage(record)
            year_data.append(result['total_percentage']['total_percentage'])
        if len(year_data) == 0:
            year_percentage = 0 
        else:   
            year_percentage = sum(year_data) / len(year_data)
        overall_response[year] = round(year_percentage,2)
    overall_response['E2'] = 40
    overall_response['E3'] = 50
    overall_response['E4'] = 60
    return overall_response
# Create User (Faculty or Student)
@router.post("/create-student", response_model=Student)
async def create_student(student_data: Student = Depends(Student)):
    # Check if the student already exists (optional)
    existing_student = await student.find_one({"id_number": student_data.student_id})
    if existing_student:
        raise HTTPException(status_code=400, detail="Student with this ID already exists.")
    # Insert the new student into the database
    student_data = student_data.model_dump()
    hash_pass = auth.get_password_hash(student_data['password'])
    student_data['password'] = hash_pass
    # print(data)
    res = await student.insert_one(student_data)
    if res.inserted_id and res.acknowledged:
        return {'message':'True'}
    else:
        return {'message':'False'}
   
@router.post("/create-faculty", response_model=Faculty)
async def create_faculty(faculty_data: Faculty = Depends(Faculty),user: dict = Depends(auth.get_current_user)):
    # Check if the faculty already exists (optional)
    if user['role']!='admin':
        return {'status_code':500,'message':'you cannot access this page'}
    existing_faculty = await faculty.find_one({"email_address": faculty_data.email_address})
    if existing_faculty:
        raise HTTPException(status_code=400, detail="Faculty member with this email already exists.")
    # Insert the new faculty member into the database
    faculty_data = faculty_data.model_dump()
    hash_pass = auth.get_password_hash(faculty_data['password'])
    faculty_data['password'] = hash_pass
    result = await faculty.insert_one(faculty_data)
    # Retrieve the newly created faculty document
    created_faculty = await faculty.find_one({"_id": result._id})
    if result.inserted_id and result.acknowledged:
        return {'message':'True'}
    else:
        return {'message':'False'}
   
# Get All Users (Faculty and Students)
def convert_objectid_to_str(documents):
    if isinstance(documents, list):
        for document in documents:
            for key, value in document.items():
                if isinstance(value, ObjectId):
                    document[key] = str(value)
    elif isinstance(documents, dict):
        for key, value in documents.items():
            if isinstance(value, ObjectId):
                documents[key] = str(value)
    return documents

@router.get('/get-faculty-emails')
async def get_all_users(user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin' or (not check_admin_email(user['email_address'])):
        return {'status_code':500,'message':'you cannot access this page'}
    faculty_data = await database.faculty.find({}).to_list(None)
    if faculty_data:
        response = []
        for faculty in faculty_data:
            faculty_name = faculty['first_name'] + " " + faculty['last_name']
            email = faculty['email_address']
            response.append({"faculty_name": faculty_name, "email": email})
        return {'status_code':200,'faculty_data':response}
    else:
        return {"message": "No faculty members found.",'status_code':404}

# Delete User (Faculty or Student)
@router.delete('/delete-user')
async def delete_user(user_type: str, identifier: str,user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin':
        return {'status_code':500,'message':'you cannot access this page'}
    if user_type == "faculty":
        res = await faculty.delete_one(
            {'email_address': identifier})  # Assuming email_address is used to identify faculty
        if res.deleted_count > 0:
            return {'message': "Faculty deleted successfully"}
    elif user_type == "student":
        res = await student.delete_one({'id_number': identifier})  # Assuming id_number is used to identify students
        if res.deleted_count > 0:
            return {'message': "Student deleted successfully"}
    raise HTTPException(status_code=404, detail="User  not found or invalid user type")

@router.put('/update-faculty')
async def update_faculty(email: str,faculty_data: Faculty = Depends(Faculty),user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin':
        return {'status_code':500,'message':'you cannot access this page'}
    existing_faculty = await faculty.find_one({'email_address': email})
    if not existing_faculty:
        raise HTTPException(status_code=404, detail="Faculty not found")

    faculty_data = faculty_data.model_dump()
    hash_pass = auth.get_password_hash(faculty_data['password'])
    faculty_data['password'] = hash_pass    
    res = await faculty.update_one(
            {'email_address': email},
            {'$set': faculty_data}
        )
    return {"message": "Faculty updated successfully"} if res.modified_count else {"message": "No changes made"}
 
 
@router.put('/update-student')
async def update_student(student_id: str,student_data: Student = Depends(Student),user: dict = Depends(auth.get_current_user)):   
    if user['role']!='admin':
        return {'status_code':500,'message':'you cannot access this page'}
    # Update logic for student
    existing_student = await student.find_one({'id_number': student_id})
    if not existing_student:
        raise HTTPException(status_code=404, detail="Student not found")
    student_data = student_data.model_dump()
    hash_pass = auth.get_password_hash(student_data['password'])
    student_data['password'] = hash_pass
    res = await student.update_one(
            {'id_number': student_id},
            {'$set': student_data}
        )
    return {"message": "Student updated successfully"} if res.modified_count else {"message": "No changes made"}

@router.get('/download-attendance')
async def get_attendance_summary(user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin' or (not check_admin_email(user['email_address'])):
        return {'status_code':500,'message':'you cannot access this page'}
    workbook = Workbook()
    E1 = workbook.active
    await create_sub_sheet(E1,"E1")
    E2 = workbook.create_sheet(title="E2")
    await create_sub_sheet(E2,"E2")
    E3 = workbook.create_sheet(title="E3")
    await create_sub_sheet(E3,"E3")
    E4 = workbook.create_sheet(title="E4")
    await create_sub_sheet(E4,"E4")
    

    excel_stream = BytesIO()
    workbook.save(excel_stream)
    excel_stream.seek(0)
    response = StreamingResponse(
        excel_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = "attachment; filename=preview_data.xlsx"

    return response


async def create_sub_sheet(sheet,year):
    sno = 1
    sheet.title = f"{year}"
    sheet.merge_cells("A1:D1") 
    sheet["A1"] = "Student Details"
    sheet["A1"].font = Font(name="Times New Roman", size=17, bold=True)
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    # List of column groups to merge
    column_groups = [
        ('E', 'N'),
        ('O', 'X'),
        ('Y', 'AH'),
        ('AI', 'AR'),
        ('AS', 'BB'),
        ('BC', 'BL'),
        ('BM', 'BV'),
        ('BW', 'CF'),
        ('CG', 'CP'),
        ('CQ', 'DA'),
        ('DB', 'DJ'),
        ('DK', 'DS')
    ]
    year_details = student_collections[year]
    prefix_exams = exam_timetable_collections[year]
    exams = await prefix_exams.find_one({})
    subjects_data = await year_details.find_one({})
    if exams is None or subjects_data is None:
        raise HTTPException(status_code=400, detail="Invalid year")
    subjects = [subject['subject_name'] for subject in subjects_data['subjects']]
    for subject , (start_col, end_col) in zip(subjects,column_groups):
        sheet.merge_cells(f"{start_col}1:{end_col}1")
        sheet[f"{start_col}1"] = f'Subject Name : {subject}'
        sheet[f"{start_col}1"].alignment = Alignment(horizontal="center", vertical="center")
        sheet[f"{start_col}1"].font = Font(name="Times New Roman", size=17, bold=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    headers = ["S No", "ID Number", "Student Name", "Section Name"]

    for col_idx, header in enumerate(headers, start=1):
        sheet.merge_cells(start_row=2, start_column=col_idx, end_row=3, end_column=col_idx)
        cell = sheet.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(name="Times New Roman", size=12, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
    
    sections = [
        {"name": f"MID-1 upto({exams['mid_exams']['MID-1']})", "subheaders": ["Number of classes Taken", "Number of classes Attended", "MID-1 Percentage"]},
        {"name": f"MID-2 upto({exams['mid_exams']['MID-1']})", "subheaders": ["Number of classes Taken", "Number of classes Attended", "MID-2 Percentage"]},
        {"name": f"MID-3 upto({exams['mid_exams']['MID-1']})", "subheaders": ["Number of classes Taken", "Number of classes Attended", "MID-3 Percentage"]},
    ]


    start_column = "E"
    for i in range(len(subjects)):  
        create_section(sheet, start_column, "Faculty Name", [])
        start_column = chr(ord(start_column) + 1)  

        for section in sections:
            create_section(sheet, start_column, section["name"], section["subheaders"])
            start_column = chr(ord(start_column) + len(section["subheaders"]))
    next_column = get_column_letter(sheet.max_column+1)
    sheet.merge_cells(f"{next_column}1:{next_column}3")
    sheet[f"{next_column}1"] = "Consolidated Percentage"
    sheet[f"{next_column}1"].font = Font(name="Times New Roman", size=17, bold=True)
    sheet[f"{next_column}1"].alignment = Alignment(horizontal="center", vertical="center",wrap_text=True)
    
    documents = await year_details.find({}).to_list(length=None)
    for document in documents:
        section = document['section_name']
        students = document['students']
        for student_id in students:
            student_details = await database.student.find_one({"id_number": student_id})
            attendance = await attendance_collections[year].find_one({"id_number": student_id})
            if attendance is None:
                continue
            if not student_details:
                continue
            student_name = student_details['first_name'] + " " + student_details['last_name']
            row = [sno, student_id, student_name, section]
            for subject in subjects:
                consolidated_attendance = 0
                row.append(attendance['attendance_report'][subject]['faculty_name'])
                mid_1_attendance = get_attendance_report(attendance, exams['mid_exams']['MID-1'],subject)
                mid_2_attendance = get_attendance_report(attendance, exams['mid_exams']['MID-2'],subject)
                mid_3_attendance = get_attendance_report(attendance, exams['mid_exams']['MID-3'],subject)
                print(mid_1_attendance,mid_2_attendance,mid_3_attendance)   
                if mid_1_attendance[0] > 0:
                    row.extend(mid_1_attendance)
                else:
                    row.extend(['', '', '']) 
                if mid_2_attendance[0] > 0:
                    row.extend(mid_2_attendance)
                else:
                    row.extend(['', '', ''])  
                if mid_3_attendance[0] > 0:
                    row.extend(mid_3_attendance)
                else:
                    row.extend(['', '', ''])
            # print(row)
            row.append(student_details['overall_attendance'])
            sheet.append(row)
            sno += 1
                
            
            # row = [sno, student_id, student_name, section]

    sheet.row_dimensions[1].height = 45
    sheet.row_dimensions[2].height = 31.5
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.border = thin_border

    
def get_attendance_report(attendance, exam_date, subject):
    attendance_report = attendance['attendance_report'][subject]['attendance']
    number_of_classes = 0
    number_of_attended = 0
    for report in attendance_report:
        if report['date'] >= exam_date:
            number_of_classes += report['number_of_periods']
            if report['status'] == 'present':
                number_of_attended += report['number_of_periods']
    
    if number_of_classes == 0:
        return [number_of_classes, number_of_attended, 0]
    else:
        return [number_of_classes, number_of_attended, (number_of_attended / number_of_classes) * 100]
    
    
    
    
    
def create_section(sheet, col_start, section_name, subheaders):
    header_font = Font(name="Times New Roman", size=12, bold=True)
    center_align = Alignment(horizontal="center", vertical="center",wrap_text=True)
    
    col_start_index = ord(col_start) - ord('A') + 1  
    if not subheaders:
        col_end_index = col_start_index
        col_start_letter = get_column_letter(col_start_index)
        col_end_letter = get_column_letter(col_end_index)
        sheet.merge_cells(f"{col_start_letter}2:{col_end_letter}3")
        sheet[f"{col_start_letter}2"] = section_name
        sheet[f"{col_start_letter}2"].font = header_font
        sheet[f"{col_start_letter}2"].alignment = center_align
    else:
        col_end_index = col_start_index + len(subheaders) - 1
        col_start_letter = get_column_letter(col_start_index)
        col_end_letter = get_column_letter(col_end_index) 
        sheet.merge_cells(f"{col_start_letter}2:{col_end_letter}2")
        sheet[f"{col_start_letter}2"] = section_name
        sheet[f"{col_start_letter}2"].font = header_font
        sheet[f"{col_start_letter}2"].alignment = center_align
    
    
    for i, subheader in enumerate(subheaders):
        col_index = col_start_index + i
        col_letter = get_column_letter(col_index)
        sheet[f"{col_letter}3"] = subheader
        sheet[f"{col_letter}3"].font = header_font
        sheet[f"{col_letter}3"].alignment = center_align


@router.delete('/delete-attendance')
async def delete_attendance(user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin' or (not check_admin_email(user['email_address'])):
        return {'status_code':500,'message':'you cannot access this page'}
    years = ['E1','E2','E3','E4']
    await database.student.update_many({}, {'$set': { 'overall_attendance': 0}})
    data = dict()
    try:
        for year in years:
            attendance_collection = attendance_collections[year]
            result = attendance_collection.delete_many({})
            data[year] = result.delete_count
        return {"message":"Attendance deleted sucessfully","data":data}
    except Exception as e:
        raise HTTPException(status_code = 500,detail=f"error while deleting the attendance collection : {str(e)}")

@router.get("/timetable/{year}")
async def get_timetable(year: str):
    """Fetch the timetable for a given academic year."""
    prefix = timetable_collections.get(year)

    timetable = await prefix.find_one({})
    if not timetable:
        raise HTTPException(status_code=404, detail="Timetable not found")

    # Convert ObjectId to string for JSON serialization
    timetable["_id"] = None  
    return {"status_code": 200, "timetable": timetable}
    
    

@router.post("/timetable")
async def modify_timetable(request: TimeTableRequest,user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin' or (not check_admin_email(user['email'])):
        return {'status_code':500,'message':'you cannot access this page'}
    year  = request.year
    prefix = timetable_collections[year]
    await prefix.delete_many({})
    inserted_result = await prefix.insert_one(request.timetable.dict())
    if inserted_result.inserted_id:
        return {"message":"Time table Sucessfully inserted" , "status_code":200}
    else:
        return {'status_code':500,'message':"Internal server error"}
    

@router.get('/visualize-attendance')
async def visualize_attendance(user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin' or (not check_admin_email(user['email'])):
        return {'status_code':500,'message':'you cannot access this page'}
    try:
        years = ['E1']
        sections = ['A', 'B', 'C', 'D', 'E']
        overall_response = {}
        total_year_attendance = []

        for year in years:
            year_data = {}
            total_students=0
            for section in sections:
                section_data = await student_collections[year].find_one({'section_name': section})
                if not section_data:
                    continue
                students = section_data.get("students", [])
                students_attendance = []
                total_students += len(students)
                for student_id in students:
                    student_attendance = await database.student.find_one({"id_number": student_id})
                    if not student_attendance:
                        continue
                    student_percentage = student_attendance.get("overall_attendance", 0)
                    students_attendance.append(student_percentage)

                if students_attendance:
                    average_percentage = sum(students_attendance) / len(students_attendance)
                else:
                    average_percentage = 0
                
                year_data[section] = average_percentage
                total_year_attendance.extend(students_attendance)

            if total_year_attendance:
                year_data["total_percentage"] = sum(total_year_attendance) / len(total_year_attendance)
            else:
                year_data["total_percentage"] = 0
            year_data["total_students"] = total_students
            overall_response[year] = year_data
        overall_response["E2"] = {"A": 50,"B": 0,"C": 50,"D": 50,"E": 50,"total_percentage": 50,"total_students": 15}
        overall_response["E3"] = {"A": 50,"B": 0,"C": 50,"D": 50,"E": 50,"total_percentage": 50,"total_students": 15}
        overall_response["E4"] = {"A": 50,"B": 0,"C": 50,"D": 50,"E": 50,"total_percentage": 50,"total_students": 15}
        return overall_response
    
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
    
@router.get('/get-subjects')
async def get_subjects(user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin' or (not check_admin_email(user['email'])):
        return {'status_code':500,'message':'you cannot access this page'}
    subjects = {}
    years = ['E1','E2','E3','E4']
    for year in years:
        subjects_data = await subjects_collections[year].find_one({'sem':'sem-2'})
        if not subjects_data:
            return {'status_code':404,'message':'subjects reocrds not found'}
        subjects[year] = subjects_data['subjects']
    
    return subjects
@router.get('/get-faculty')
async def get_faculty(user: dict = Depends(auth.get_current_user)):
    if user['role']!='admin' or (not check_admin_email(user['email'])):
        return {'status_code':500,'message':'you cannot access this page'}
    try:
        response = []
        years = ['E1']
        for year in years:
            assignments_data = await timetable_collections[year].find_one({'type': 'assignments'})
            if not assignments_data:
                raise HTTPException(status_code=404, detail="Assignments data not found")
            
            faculties_data = await database.faculty.find({}).to_list(None)
            faculty_dict = {}
            
            for faculty in faculties_data:
                faculty_email = faculty.get("email_address")
                faculty_info = {
                    "first_name": faculty.get("first_name"),
                    "last_name": faculty.get("last_name"),
                    "email": faculty_email,
                    "phone_number": faculty.get("phone_number"),
                    "department": faculty.get("department"),
                    "designation": faculty.get("designation"),
                    "qualification": faculty.get("qualification"),
                    "subjects": {}
                }
                faculty_dict[faculty_email] = faculty_info

            for subject, faculty_list in assignments_data.get("subjects", {}).items():
                for faculty_entry in faculty_list:
                    faculty_email = faculty_entry["faculty_username"]
                    if faculty_email in faculty_dict:
                        faculty_dict[faculty_email]["subjects"][year] = {
                            "subject": subject,
                            "sections": faculty_entry["sec"]
                        }
            
            response = list(faculty_dict.values())
        return {"message": "Success", "faculty_details": response}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")
